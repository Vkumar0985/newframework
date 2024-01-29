import openpyxl


class Excel:
    @staticmethod
    def get_cell_data(path, sheet, row, col):
        try:
            wb = openpyxl.load_workbook(path)
            v = wb[sheet].cell(row, col).value
            return v
        except:
            return ""

    @staticmethod
    def write_cell_data(path, sheet, row, col, val):
        try:
            wb = openpyxl.load_workbook(path)
            # wb.active()
            wb[sheet].cell(row, col).value = val
            wb.save(path)
            return True
        except:
            return False
